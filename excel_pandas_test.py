import pandas as pd
import openpyxl
import os
import sys
import datetime

def unmerge_cells(input_file):
    """Excelファイルの結合セルを解除し、値を全てのセルに複製する"""
    # 一時ファイル名を作成
    base, ext = os.path.splitext(input_file)
    output_file = f"{base}_unmerged{ext}"
    
    # Excelファイルを読み込む。
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # 結合セルの情報を取得
    merged_cells = list(ws.merged_cells.ranges)
    
    # 結合セルを解除し、値を全てのセルに複製
    for merged_range in merged_cells:
        # 結合セルの左上のセルから値を取得
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value
        
        # 結合を解除
        ws.unmerge_cells(str(merged_range))
        
        # 解除したセル全てに値をコピー
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col).value = value
    
    # 変更を保存
    wb.save(output_file)
    print(f"結合セルを解除したファイルを {output_file} に保存しました。")
    return output_file

def merge_multirow_headers(df):
    # 最初の行を取得（1行目のみを使用）
    row1 = df.iloc[0]
    
    # 1行目の値だけを使ってカラム名を作成
    new_columns = []
    for i in range(len(row1)):
        # 値が空でないことを確認
        if not pd.isna(row1.iloc[i]) and row1.iloc[i] != '':
            new_columns.append(str(row1.iloc[i]))
        else:
            # 1行目が空の場合は列番号をプレースホルダーとして使用
            new_columns.append(f"Column_{i+1}")
    
    # 新しいカラム名を設定
    df.columns = new_columns
    
    # 最初の1行を削除
    df = df.iloc[1:].reset_index(drop=True)
    
    return df
    
def main():
    # コマンドライン引数をチェック
    if len(sys.argv) < 2:
        sys.exit(1)
    
    # 入力ファイル名をコマンドライン引数から取得
    input_file = sys.argv[1]
    
    # 1. まず結合セルを解除
    unmerged_file = unmerge_cells(input_file)
    
    # 2. 結合セルが解除されたファイルを読み込む
    df = pd.read_excel(unmerged_file)
    
    # 3. データフレームを処理
    df_merged = merge_multirow_headers(df)
    
    # 4. 結果の確認
    print(df_merged.head())
    
    # 5. 現在の時間をYYYYMMDD-hh.mm形式で取得
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y%m%d-%H%M%S")
    
    # 6. 最終的な処理結果をExcelファイルとして出力（ファイル名に時間を追加）
    base, ext = os.path.splitext(input_file)
    output_file = f"{base}_処理済_{timestamp}{ext}"
    df_merged.to_excel(output_file, index=False)
    print(f"処理結果を {output_file} に保存しました。")

if __name__ == "__main__":
    main()
